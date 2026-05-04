import streamlit as st
import sys
import os
import pandas as pd
from io import BytesIO
from datetime import date, timedelta

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from database import init_db, get_report_data, get_stores, get_users, get_states

init_db()

st.title("📑 Reports")
st.markdown("---")

# ── Filters ────────────────────────────────────────────────────────────────────
st.subheader("Filters")
col1, col2, col3, col4 = st.columns(4)
with col1:
    start_date = st.date_input("Start Date", value=date.today() - timedelta(days=30))
with col2:
    end_date = st.date_input("End Date", value=date.today())

stores  = get_stores()
store_map = {"All Stores": None}
store_map.update({s['store_name']: s['id'] for s in stores})

states = get_states()
state_opts = ["All States"] + states

techs = get_users()
tech_map = {"All Technicians": None}
tech_map.update({t['name']: t['id'] for t in techs})

with col3:
    sel_store = st.selectbox("Store", list(store_map.keys()))
with col4:
    sel_state = st.selectbox("State", state_opts)

col1, col2 = st.columns([1, 3])
with col1:
    sel_tech = st.selectbox("Technician / Vendor", list(tech_map.keys()))

if st.button("Generate Report", type="primary"):
    if start_date > end_date:
        st.error("Start date must be before end date.")
        st.stop()

    rows = get_report_data(
        start_date=start_date.isoformat(),
        end_date=end_date.isoformat(),
        store_id=store_map.get(sel_store),
        state=None if sel_state == "All States" else sel_state,
        tech_id=tech_map.get(sel_tech),
    )

    if not rows:
        st.info("No records found for the selected filters.")
    else:
        df = pd.DataFrame(rows, columns=[
            'Store', 'State', 'Date', 'Technician',
            'AC Number', 'Serial Number', 'Capacity (Ton)', 'AC Type',
            'AC Image Path', 'Serial Image Path', 'Status'
        ])

        st.markdown(f"**{len(df)} record(s) found**")
        st.dataframe(df.drop(columns=['AC Image Path', 'Serial Image Path']),
                     use_container_width=True, height=420)

        def build_excel(dataframe):
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                dataframe.to_excel(writer, sheet_name='PMS Report', index=False)
                ws = writer.sheets['PMS Report']
                from openpyxl.styles import PatternFill, Font, Alignment
                hfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                for cell in ws[1]:
                    cell.fill = hfill
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal='center')
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = min(
                        max(len(str(c.value or '')) for c in col) + 3, 45
                    )
            return buf.getvalue()

        st.download_button(
            "📥 Download Excel Report",
            data=build_excel(df),
            file_name=f"PMS_Report_{start_date}_{end_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.markdown("---")
        st.subheader("Summary")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Records",       len(df))
        c2.metric("Unique Stores",       df['Store'].nunique())
        c3.metric("Unique Technicians",  df['Technician'].nunique())
        c4.metric("States Covered",      df['State'].nunique())