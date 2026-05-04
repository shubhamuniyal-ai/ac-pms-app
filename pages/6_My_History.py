import streamlit as st
import sys
import os
import pandas as pd
from io import BytesIO

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from database import init_db, get_vendor_history, get_session_ac_entries

init_db()

user = st.session_state.get('user', {})
st.title("🕐 My PMS History")
st.caption(f"Showing all PMS submissions by **{user.get('name','')}**")
st.markdown("---")

sessions = get_vendor_history(user['id'])

if not sessions:
    st.info("You haven't submitted any PMS entries yet. Go to **PMS Entry** to get started.")
    st.stop()

# ── Summary cards ──────────────────────────────────────────────────────────────
total_ac = sum(s['ac_count'] for s in sessions)
unique_stores = len({s['store_name'] for s in sessions})

c1, c2, c3 = st.columns(3)
c1.metric("Total Sessions", len(sessions))
c2.metric("Total ACs Recorded", total_ac)
c3.metric("Stores Covered", unique_stores)

st.markdown("---")

# ── Filter ─────────────────────────────────────────────────────────────────────
store_list = sorted({s['store_name'] for s in sessions})
sel_store  = st.selectbox("Filter by Store", ["All Stores"] + store_list)

if sel_store != "All Stores":
    sessions = [s for s in sessions if s['store_name'] == sel_store]

# ── Session list ───────────────────────────────────────────────────────────────
st.subheader(f"{len(sessions)} Session(s)")

for idx, sess in enumerate(sessions):
    label = (
        f"📋  {sess['entry_date']}  ·  {sess['store_name']}  ·  "
        f"{sess['state']}  ·  {sess['ac_type']}  ·  {sess['ac_count']} AC(s)"
    )
    with st.expander(label, expanded=(idx == 0)):
        info_col, badge_col = st.columns([4, 1])
        with info_col:
            st.write(f"**Store:** {sess['store_name']}  |  **State:** {sess['state']}")
            st.write(f"**Date:** {sess['entry_date']}  |  **AC Type:** {sess['ac_type']}")
            st.write(f"**Submitted:** {sess['created_at']}")
        with badge_col:
            color = "#22c55e" if sess['status'] == 'Completed' else "#f59e0b"
            st.markdown(
                f'<div style="background:{color};color:#fff;padding:6px 12px;'
                f'border-radius:20px;text-align:center;font-size:.85rem;font-weight:600">'
                f'{sess["status"]}</div>',
                unsafe_allow_html=True
            )

        entries = get_session_ac_entries(sess['session_id'])
        if entries:
            st.markdown("**AC Entries:**")
            for i, e in enumerate(entries):
                with st.container():
                    ec1, ec2, ec3, ec4 = st.columns([2, 2, 1, 2])
                    with ec1:
                        st.write(f"**AC #{i+1}** — `{e['ac_number']}`")
                    with ec2:
                        st.write(f"Serial: `{e['serial_number']}`")
                    with ec3:
                        st.write(f"{e['capacity']} Ton")
                    with ec4:
                        if e.get('ac_number_image') and os.path.exists(e['ac_number_image']):
                            with open(e['ac_number_image'], 'rb') as f:
                                st.download_button(
                                    "📷 AC Photo",
                                    data=f.read(),
                                    file_name=os.path.basename(e['ac_number_image']),
                                    key=f"dl_ac_{e['id']}"
                                )
                        if e.get('serial_number_image') and os.path.exists(e['serial_number_image']):
                            with open(e['serial_number_image'], 'rb') as f:
                                st.download_button(
                                    "📷 Serial Photo",
                                    data=f.read(),
                                    file_name=os.path.basename(e['serial_number_image']),
                                    key=f"dl_sr_{e['id']}"
                                )
                    st.divider()

# ── Export my history ──────────────────────────────────────────────────────────
st.markdown("---")
if st.button("📥 Export My History to Excel"):
    rows = []
    for sess in sessions:
        entries = get_session_ac_entries(sess['session_id'])
        for e in entries:
            rows.append({
                'Date': sess['entry_date'],
                'Store': sess['store_name'],
                'State': sess['state'],
                'AC Type': sess['ac_type'],
                'AC Number': e['ac_number'],
                'Serial Number': e['serial_number'],
                'Capacity (Ton)': e['capacity'],
                'Status': sess['status'],
                'Submitted At': sess['created_at'],
            })

    if rows:
        df = pd.DataFrame(rows)
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as w:
            df.to_excel(w, index=False, sheet_name='My History')
            ws = w.sheets['My History']
            from openpyxl.styles import PatternFill, Font, Alignment
            hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or '')) for c in col) + 3, 40
                )
        st.download_button(
            "⬇️ Download Excel",
            data=buf.getvalue(),
            file_name=f"my_pms_history_{user.get('name','').replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )